#!/usr/bin/env python3
"""
Build the Marketshare SQLite database from scratch and import the two
source spreadsheets:
  - jog_fleet_combined.xlsx  -> boat-level fleet register + JOG entry lists
  - IRC_Solent_Report.xlsx   -> aggregate historical class counts per regatta/year

Usage: python3 build_db.py <jog_fleet_combined.xlsx> <IRC_Solent_Report.xlsx> <out.db>
"""
import sys
import re
import sqlite3
import datetime
import openpyxl

SCHEMA_PATH = "/home/claude/marketshare/db/schema.sql"

# Sailmakers seen across the sheets. North Sails is "us".
SAILMAKERS = [
    ("North Sails", 1), ("Doyle", 0), ("Ullman", 0), ("Quantum", 0),
    ("Sanders", 0), ("Partial", 0), ("Unknown", 0), ("Other", 0),
    ("GP", 0), ("UK", 0),
]
# Normalize raw sailmaker text -> canonical name in SAILMAKERS
SAILMAKER_ALIASES = {
    "north": "North Sails", "north sails": "North Sails",
    "doyle": "Doyle", "ullman": "Ullman", "quantum": "Quantum",
    "sanders": "Sanders", "partial": "Partial", "unknown": "Unknown",
    "other": "Other", "gp": "GP", "uk": "UK",
}


def norm(s):
    if s is None:
        return None
    s = str(s).strip()
    return s if s else None


def norm_upper(s):
    s = norm(s)
    return s.upper() if s else None


def get_or_create_sailmaker(cur, raw_name):
    n = norm(raw_name)
    if not n:
        return None
    canonical = SAILMAKER_ALIASES.get(n.lower(), n)
    cur.execute("SELECT id FROM sailmakers WHERE name = ?", (canonical,))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute("INSERT INTO sailmakers (name, is_us) VALUES (?, 0)", (canonical,))
    return cur.lastrowid


def get_or_create_owner(cur, raw_name):
    n = norm_upper(raw_name)
    if not n:
        return None
    cur.execute("SELECT id FROM owners WHERE name = ?", (n,))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute("INSERT INTO owners (name) VALUES (?)", (n,))
    return cur.lastrowid


def get_or_create_boat(cur, sail_no, boat_name, boat_type=None, tcc=None):
    sail_no = norm_upper(sail_no)
    boat_name = norm_upper(boat_name)
    if not sail_no and not boat_name:
        return None
    boat_id = None
    if sail_no:
        cur.execute("SELECT id FROM boats WHERE sail_no = ?", (sail_no,))
        row = cur.fetchone()
        if row:
            boat_id = row[0]
    if boat_id is None and sail_no is None and boat_name:
        # fallback match by name only (rare boats without a clean sail no)
        cur.execute("SELECT id FROM boats WHERE boat_name = ? AND sail_no IS NULL", (boat_name,))
        row = cur.fetchone()
        if row:
            boat_id = row[0]
    if boat_id is None:
        cur.execute(
            "INSERT INTO boats (sail_no, boat_name, boat_type, tcc) VALUES (?, ?, ?, ?)",
            (sail_no, boat_name, norm(boat_type), tcc),
        )
        boat_id = cur.lastrowid
        cur.execute(
            "INSERT INTO boat_crm (boat_id) VALUES (?)", (boat_id,)
        )
    else:
        # refresh "most recently observed" fields
        cur.execute(
            "UPDATE boats SET boat_name = COALESCE(?, boat_name), "
            "boat_type = COALESCE(?, boat_type), tcc = COALESCE(?, tcc), "
            "updated_at = datetime('now') WHERE id = ?",
            (boat_name, norm(boat_type), tcc, boat_id),
        )
    return boat_id


def get_or_create_regatta(cur, name, category, region="Solent"):
    cur.execute("SELECT id FROM regattas WHERE name = ?", (name,))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute(
        "INSERT INTO regattas (name, category, region) VALUES (?, ?, ?)",
        (name, category, region),
    )
    return cur.lastrowid


def get_or_create_event(cur, regatta_id, season_year, notes=None, source_url=None):
    cur.execute(
        "SELECT id FROM events WHERE regatta_id = ? AND season_year = ?",
        (regatta_id, season_year),
    )
    row = cur.fetchone()
    if row:
        if notes:
            cur.execute("UPDATE events SET notes = COALESCE(notes, ?) WHERE id = ?", (notes, row[0]))
        return row[0]
    cur.execute(
        "INSERT INTO events (regatta_id, season_year, notes, source_url) VALUES (?, ?, ?, ?)",
        (regatta_id, season_year, notes, source_url),
    )
    return cur.lastrowid


def create_race(cur, event_id, race_name, race_number=None, status=None):
    cur.execute(
        "INSERT INTO races (event_id, race_name, race_number, status) VALUES (?, ?, ?, ?)",
        (event_id, race_name, race_number, status),
    )
    return cur.lastrowid


# ---------------------------------------------------------------------------
# jog_fleet_combined.xlsx
# ---------------------------------------------------------------------------

JOG_SEASON_YEAR = 2026  # ASSUMPTION: current-season fleet register, no year in sheet.
# Flagged clearly in the README / dashboard footer for the user to correct if wrong.

def import_jog_fleet(cur, path):
    wb = openpyxl.load_workbook(path, data_only=True)

    # 1) Master fleet register ('Boats' tab) -> boats + boat_crm + sailmaker_history
    ws = wb["Boats"]
    headers = [norm(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}
    n_boats = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        sail_no = row[col["Sail No"]]
        if not norm(sail_no):
            continue
        boat_name = row[col["Boat Name"]]
        tcc = row[col["TCC"]]
        boat_type = row[col["Boat Type"]]
        skipper = row[col.get("Skipper/s")] if "Skipper/s" in col else None
        owner_name = row[col.get("Owner")] if "Owner" in col else None
        sailmaker_raw = row[col.get("Sailmaker")] if "Sailmaker" in col else None
        lead_rep = row[col.get("Lead Rep")] if "Lead Rep" in col else None
        contacted_by = row[col.get("Contacted by")] if "Contacted by" in col else None
        in_cs = row[col.get("In CS")] if "In CS" in col else None
        tag = row[col.get("Tag")] if "Tag" in col else None

        boat_id = get_or_create_boat(cur, sail_no, boat_name, boat_type, tcc)
        owner_id = get_or_create_owner(cur, owner_name)
        if owner_id:
            cur.execute("UPDATE boats SET current_owner_id = ? WHERE id = ?", (owner_id, boat_id))

        cur.execute(
            "UPDATE boat_crm SET lead_rep = ?, contacted_by = ?, in_cs = ?, tag = ?, "
            "last_updated = datetime('now') WHERE boat_id = ?",
            (norm(lead_rep), norm(contacted_by),
             (1 if norm(in_cs) else 0) if in_cs is not None else None,
             norm(tag), boat_id),
        )

        sm_id = get_or_create_sailmaker(cur, sailmaker_raw)
        if sm_id:
            cur.execute(
                "INSERT INTO boat_sailmaker_history (boat_id, sailmaker_id, source, confidence) "
                "VALUES (?, ?, 'manual:jog_fleet_combined:boats', 'manual')",
                (boat_id, sm_id),
            )
        n_boats += 1
    print(f"  Boats master register: {n_boats} boats")

    # 2) Per-race entry-list sheets
    event_sheets = [
        ("JOG Lonely Tower", "JOG Lonely Tower"),
        ("Easter Challenge", "JOG Easter Challenge"),
        ("Jog Cherbourg", "JOG Cherbourg"),
    ]
    # Sheet3 backfills sailmaker for Jog Cherbourg (that sheet's Sailmaker col is empty)
    sheet3_sailmaker_by_sailno = {}
    if "Sheet3" in wb.sheetnames:
        ws3 = wb["Sheet3"]
        for row in ws3.iter_rows(min_row=2, values_only=True):
            sail_no = norm_upper(row[0]) if row and row[0] else None
            sm = row[6] if row and len(row) > 6 else None
            if sail_no and norm(sm):
                sheet3_sailmaker_by_sailno[sail_no] = sm

    for sheet_name, regatta_name in event_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [norm(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers) if h}
        sail_no_key = "Sail No" if "Sail No" in col else "Sail no."
        regatta_id = get_or_create_regatta(cur, regatta_name, "JOG")
        event_id = get_or_create_event(cur, regatta_id, JOG_SEASON_YEAR,
                                        notes="Season year assumed from upload date; confirm/correct.")
        race_id = create_race(cur, event_id, regatta_name, status="entered")

        n_entries = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            sail_no = row[col[sail_no_key]]
            if not norm(sail_no):
                continue
            boat_name = row[col.get("Boat Name")]
            tcc = row[col.get("TCC")]
            boat_type = row[col.get("Boat Type")]
            skipper = row[col.get("Skipper/s")] if "Skipper/s" in col else None
            owner_name = row[col.get("Owner")] if "Owner" in col else None
            sailmaker_raw = row[col.get("Sailmaker")] if "Sailmaker" in col else None
            lead_rep = row[col.get("Lead Rep")] if "Lead Rep" in col else None
            contacted_by = row[col.get("Contacted by")] if "Contacted by" in col else None
            in_cs = row[col.get("In CS")] if "In CS" in col else None
            tag = row[col.get("Tag")] if "Tag" in col else None

            if not norm(sailmaker_raw):
                sailmaker_raw = sheet3_sailmaker_by_sailno.get(norm_upper(sail_no))

            boat_id = get_or_create_boat(cur, sail_no, boat_name, boat_type, tcc)
            owner_id = get_or_create_owner(cur, owner_name)
            sm_id = get_or_create_sailmaker(cur, sailmaker_raw)

            cur.execute(
                "INSERT OR REPLACE INTO race_entries "
                "(id, race_id, boat_id, sail_no_used, boat_name_used, boat_type_used, tcc, "
                " owner_id, owner_name_used, skipper_name_used, sailmaker_id, status, "
                " lead_rep, contacted_by, in_cs, tag, source) "
                "VALUES ("
                "  (SELECT id FROM race_entries WHERE race_id = ? AND boat_id = ?),"
                "  ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'entered', ?, ?, ?, ?, ?)",
                (race_id, boat_id,
                 race_id, boat_id, norm_upper(sail_no), norm_upper(boat_name), norm(boat_type), tcc,
                 owner_id, norm_upper(owner_name), norm_upper(skipper), sm_id,
                 norm(lead_rep), norm(contacted_by),
                 (1 if norm(in_cs) else 0) if in_cs is not None else None,
                 norm(tag), "manual:jog_fleet_combined"),
            )
            n_entries += 1
        print(f"  {sheet_name}: {n_entries} entries -> regatta '{regatta_name}' ({JOG_SEASON_YEAR})")


# ---------------------------------------------------------------------------
# IRC_Solent_Report.xlsx  (aggregate, non-boat-level historical counts)
# ---------------------------------------------------------------------------

YEAR_RE = re.compile(r"^\s*(\d{4})")


def year_from_header(v):
    if v is None:
        return None
    m = YEAR_RE.match(str(v))
    return int(m.group(1)) if m else None


def scan_sheet_blocks(ws, class_col=1, max_row=200):
    """Scan an entire sheet for '<title row>, Class, <year>, <year>, ...' blocks
    (there can be several stacked vertically, e.g. RSYC's May/June/July/September
    regattas). Returns a list of dicts: {title, header_row, year_cols, data_rows}
    where year_cols is {col_idx: year} and data_rows is [(class_label, row_idx)].
    """
    blocks = []
    r = 1
    while r <= max_row:
        val = ws.cell(row=r, column=class_col).value
        if val is not None and str(val).strip().lower() == "class":
            title_val = ws.cell(row=r - 1, column=class_col).value
            title = norm(title_val) if title_val else None
            year_cols = {}
            c = class_col + 1
            while True:
                yr = year_from_header(ws.cell(row=r, column=c).value)
                if yr is None:
                    break
                year_cols[c] = yr
                c += 1
            data_rows = []
            rr = r + 1
            while rr <= max_row:
                cval = norm(ws.cell(row=rr, column=class_col).value)
                if cval is None:
                    break
                low = cval.lower()
                if low.startswith("comment") or low.startswith("link"):
                    break
                data_rows.append((cval, rr))
                rr += 1
                if low == "total":
                    break
            blocks.append({"title": title, "header_row": r, "year_cols": year_cols, "data_rows": data_rows})
            r = rr
        else:
            r += 1
    return blocks


def import_class_block(cur, ws, fallback_name, category, class_col=1, max_row=200,
                        notes_by_year=None, title_prefix="", title_overrides=None):
    """Parse every 'Class | year | year | ...' block in a sheet (there may be
    several stacked vertically, each its own race/regatta with its own title
    row directly above 'Class') and load them as event_class_counts.

    Each block's own title (row above 'Class') becomes its regatta name,
    after applying title_overrides (case-insensitive raw title -> final name)
    and title_prefix. Blocks with no title row (source data incomplete) fall
    back to "<fallback_name> (extra classes, row N)" so nothing collides or
    is silently dropped - these need a manual rename once reviewed.
    """
    title_overrides = {k.lower(): v for k, v in (title_overrides or {}).items()}
    blocks = scan_sheet_blocks(ws, class_col=class_col, max_row=max_row)
    total_cells = 0
    total_events = set()
    for block in blocks:
        title = block["title"]
        if title and title.lower() in title_overrides:
            name = title_overrides[title.lower()]
        elif title:
            name = f"{title_prefix}{title}"
        else:
            hint = block["data_rows"][0][0] if block["data_rows"] else f"row {block['header_row']}"
            name = f"{fallback_name} — {hint} group"
        regatta_id = get_or_create_regatta(cur, name, category)
        events_created = {}
        for class_label, rr in block["data_rows"]:
            for c, yr in block["year_cols"].items():
                count_val = ws.cell(row=rr, column=c).value
                if count_val in (None, "-", "Cancelled", "Included in IRC"):
                    continue
                try:
                    count_int = int(round(float(count_val)))
                except (ValueError, TypeError):
                    continue
                if yr not in events_created:
                    note = (notes_by_year or {}).get(yr)
                    events_created[yr] = get_or_create_event(cur, regatta_id, yr, notes=note)
                event_id = events_created[yr]
                cur.execute(
                    "INSERT OR REPLACE INTO event_class_counts "
                    "(id, event_id, class_label, entry_count, source) VALUES ("
                    " (SELECT id FROM event_class_counts WHERE event_id = ? AND class_label = ? AND source = ?),"
                    " ?, ?, ?, ?)",
                    (event_id, class_label, "manual:irc_solent_report",
                     event_id, class_label, count_int, "manual:irc_solent_report"),
                )
                total_cells += 1
        total_events |= set(events_created.values())
        print(f"  {name}: {sum(len(b['year_cols']) for b in [block])} yrs x {len(block['data_rows'])} classes")
    print(f"  -> {len(blocks)} block(s), {total_cells} class/year count cells, {len(total_events)} event-years total")


def import_irc_solent_report(cur, path):
    wb = openpyxl.load_workbook(path, data_only=True)

    # RORC Inshore: block 1 is mislabeled "Easter challenge" in the source
    # (copy/paste artifact) - it's actually the main Inshore Series classes.
    import_class_block(cur, wb["RORC Inshore"], "RORC Inshore Series", "RORC",
                        title_overrides={"easter challenge": "RORC Inshore Series"})

    # RSYC: 4 stacked blocks (May/June/July/September Regatta) -> 4 distinct regattas
    import_class_block(cur, wb["RSYC"], "RSYC Regatta", "Club", title_prefix="RSYC ")

    # Cowes Week and RTI sheet actually contains two distinct races stacked: RTI and Cowes Week
    import_class_block(cur, wb["Cowes Week and RTI"], "Cowes Week / RTI", "Club",
                        title_overrides={"rti": "Round the Island Race", "cowes week": "Cowes Week"})

    # Warsash: Spring Series and Spring Championships are two different events
    import_class_block(cur, wb["Warsash Spring Series"], "Warsash Series", "Club",
                        title_overrides={"spring series": "Warsash Spring Series",
                                          "spring championships": "Warsash Spring Championships"})

    # ORC Worlds & ORC Euros - year headers have venue in parens, e.g. "2025 (Tallin)"
    ws_orc = wb["ORC "]
    venue_by_year = {}
    for c in range(2, 12):
        val = ws_orc.cell(row=2, column=c).value
        yr = year_from_header(val)
        if yr and val:
            m = re.search(r"\(([^)]+)\)", str(val))
            if m:
                venue_by_year[yr] = f"Venue: {m.group(1)}"
    import_class_block(cur, ws_orc, "ORC Championship", "Championship", notes_by_year=venue_by_year,
                        title_overrides={"orc worlds": "ORC Worlds", "orc euros": "ORC Europeans"})

    # Offshore sheet contains three distinct RORC offshore races
    import_class_block(cur, wb["Offshore"], "RORC Offshore Race", "RORC",
                        title_overrides={"cervantes": "RORC Cervantes Trophy (Offshore)",
                                          "myth of malham": "RORC Myth of Malham",
                                          "rolex fastnet race": "Rolex Fastnet Race"})

    # Totals sheet, first block only = Cowes Week overall totals (no class breakdown)
    ws_tot = wb["Totals"]
    regatta_id = get_or_create_regatta(cur, "Cowes Week", "Club")
    year_cols = {}
    for c in range(2, 9):
        yr = year_from_header(ws_tot.cell(row=2, column=c).value)
        if yr:
            year_cols[c] = yr
    n = 0
    for c, yr in year_cols.items():
        val = ws_tot.cell(row=3, column=c).value
        if val is None:
            continue
        try:
            count_int = int(round(float(val)))
        except (ValueError, TypeError):
            continue
        event_id = get_or_create_event(cur, regatta_id, yr)
        cur.execute(
            "INSERT OR REPLACE INTO event_class_counts "
            "(id, event_id, class_label, entry_count, source) VALUES ("
            " (SELECT id FROM event_class_counts WHERE event_id = ? AND class_label = ? AND source = ?),"
            " ?, ?, ?, ?)",
            (event_id, "Total", "manual:irc_solent_report",
             event_id, "Total", count_int, "manual:irc_solent_report"),
        )
        n += 1
    print(f"  Cowes Week (totals only): {n} year totals")


def main():
    if len(sys.argv) != 4:
        print(__doc__)
        sys.exit(1)
    jog_path, irc_path, out_path = sys.argv[1:4]

    conn = sqlite3.connect(out_path)
    cur = conn.cursor()
    with open(SCHEMA_PATH) as f:
        cur.executescript(f.read())

    for name, is_us in SAILMAKERS:
        cur.execute("INSERT OR IGNORE INTO sailmakers (name, is_us) VALUES (?, ?)", (name, is_us))

    print("Importing jog_fleet_combined.xlsx ...")
    import_jog_fleet(cur, jog_path)

    print("Importing IRC_Solent_Report.xlsx ...")
    import_irc_solent_report(cur, irc_path)

    conn.commit()

    cur.execute("SELECT count(*) FROM boats")
    n_boats = cur.fetchone()[0]
    cur.execute("SELECT count(*) FROM race_entries")
    n_entries = cur.fetchone()[0]
    cur.execute("SELECT count(*) FROM event_class_counts")
    n_counts = cur.fetchone()[0]
    print(f"\nDone. {n_boats} boats, {n_entries} race entries, {n_counts} aggregate class-count rows.")
    print(f"Built at {datetime.datetime.now().isoformat()} -> {out_path}")
    conn.close()


if __name__ == "__main__":
    main()
