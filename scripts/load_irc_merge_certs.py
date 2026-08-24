#!/usr/bin/env python3
"""
Load TCC from the richer 'Merge_IRC_SailData_with_TCC_Listing_years.xlsm'
workbook (national IRC cert listing, current + Club_Listing_2026 with an
explicit Endorsed flag, plus year-by-year ALL_YYYY sheets 2020-2026) into
the database as the authoritative TCC source, overriding load_irc_certs.py.

Per sail number, picks the best row: Club_Listing 'Endorsed'=='E' first,
then most recent Cert Year, then most recent Issue Date. Falls back to the
most recent ALL_YYYY sheet for boats not in Club_Listing at all.

Usage:
  python3 load_irc_merge_certs.py <db.sqlite> <Merge_IRC_SailData...xlsm>
"""
import sys
import argparse
import sqlite3
import datetime
import openpyxl

sys.path.insert(0, str(__import__("pathlib").Path(__file__).resolve().parent))
from build_db import norm


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("db")
    p.add_argument("xlsm_file")
    return p.parse_args()


def main():
    args = parse_args()
    conn = sqlite3.connect(args.db)
    cur = conn.cursor()

    wb = openpyxl.load_workbook(args.xlsm_file, read_only=True, data_only=True, keep_vba=False)

    # ---- Club_Listing (current, has an explicit Endorsed flag) ----
    club_ws = wb["Club_Listing_2026"]
    club_header = next(club_ws.iter_rows(min_row=1, max_row=1, values_only=True))
    ci = {h: i for i, h in enumerate(club_header)}
    by_sailno = {}
    for row in club_ws.iter_rows(min_row=2, values_only=True):
        sail_no = norm(row[ci["Sail No"]])
        if not sail_no:
            continue
        tcc = row[ci["TCC"]]
        if tcc in (None, "#N/A") or not isinstance(tcc, (int, float)):
            continue
        endorsed = row[ci["Endorsed"]] == "E"
        cert_year = row[ci["Cert Year"]] or 0
        issue_date = row[ci["Issue Date"]]
        key = (1 if endorsed else 0, cert_year, issue_date or datetime.datetime.min)
        existing = by_sailno.get(sail_no)
        if existing is None or key > existing[0]:
            by_sailno[sail_no] = (key, float(tcc))

    club_matched = {k: v[1] for k, v in by_sailno.items()}

    # ---- fall back to the most recent ALL_YYYY sheet for anything else ----
    fallback = {}
    for year_sheet in ["ALL_2026", "ALL_2025", "ALL_2024", "ALL_2023", "ALL_2022", "ALL_2021", "ALL_2020"]:
        ws = wb[year_sheet]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        ai = {h: i for i, h in enumerate(header)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            sail_no = norm(row[ai["Sail no"]])
            if not sail_no or sail_no in fallback or sail_no in club_matched:
                continue
            tcc = row[ai["TCC"]]
            if tcc in (None, "#N/A") or not isinstance(tcc, (int, float)):
                continue
            fallback[sail_no] = float(tcc)

    combined = {**fallback, **club_matched}  # club-listing wins on overlap

    updated = 0
    not_found = 0
    for sail_no, tcc in combined.items():
        cur.execute("SELECT id, tcc FROM boats WHERE sail_no = ?", (sail_no,))
        boat_row = cur.fetchone()
        if not boat_row:
            not_found += 1
            continue
        boat_id, old_tcc = boat_row
        if old_tcc == tcc:
            continue
        cur.execute("UPDATE boats SET tcc = ?, updated_at = ? WHERE id = ?",
                    (tcc, datetime.datetime.now().isoformat(), boat_id))
        updated += 1

    conn.commit()
    print(f"Matched {len(combined)} boats ({len(club_matched)} from Club_Listing endorsed/latest cert, "
          f"{len(combined)-len(club_matched)} from ALL_YYYY fallback). Updated TCC for {updated} boat(s). "
          f"{not_found} sail number(s) not in the DB.")
    conn.close()


if __name__ == "__main__":
    main()
