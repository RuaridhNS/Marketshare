#!/usr/bin/env python3
"""
Load a filled-in templates/manual_entry_template.xlsx (the 'Entries' sheet)
into the Marketshare database. One row per boat per race; rows sharing the
same (Regatta, SeasonYear, RaceName) become one race with many entries.

Usage:
  python3 load_manual_template.py <db.sqlite> <filled_template.xlsx> [--category CATEGORY]

--category sets the regatta category ('JOG', 'RORC', 'Club', 'Championship')
for any NEW regatta created by this load; existing regattas keep their
current category. Defaults to 'Manual'.
"""
import sys
import argparse
import sqlite3
import openpyxl

sys.path.insert(0, "/home/claude/marketshare/scripts")
from build_db import (get_or_create_sailmaker, get_or_create_owner, get_or_create_boat,
                       get_or_create_regatta, get_or_create_event, norm, norm_upper)


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("db")
    p.add_argument("xlsx_file")
    p.add_argument("--category", default="Manual")
    return p.parse_args()


def status_from(status_raw, position, corrected_time):
    s = norm(status_raw)
    if s:
        return s.lower()
    if position or corrected_time:
        return "finished"
    return "entered"


def get_or_create_race(cur, event_id, race_name):
    cur.execute("SELECT id FROM races WHERE event_id = ? AND race_name = ?", (event_id, race_name))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute("INSERT INTO races (event_id, race_name, status) VALUES (?, ?, 'confirmed')",
                (event_id, race_name))
    return cur.lastrowid


def main():
    args = parse_args()
    conn = sqlite3.connect(args.db)
    cur = conn.cursor()

    wb = openpyxl.load_workbook(args.xlsx_file, data_only=True)
    if "Entries" not in wb.sheetnames:
        raise SystemExit("No 'Entries' sheet found in this workbook.")
    ws = wb["Entries"]
    headers = [norm(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}
    required = ["Regatta", "SeasonYear", "RaceName", "SailNo"]
    missing = [r for r in required if r not in col]
    if missing:
        raise SystemExit(f"Template missing required column(s): {missing}")

    race_cache = {}  # (regatta, year, race_name) -> race_id
    source_urls = {}  # race key -> url
    n_rows = 0
    n_races = set()
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        def get(name):
            i = col.get(name)
            return row[i] if i is not None and i < len(row) else None

        regatta = norm(get("Regatta"))
        season_year = get("SeasonYear")
        race_name = norm(get("RaceName"))
        sail_no = get("SailNo")
        notes = norm(get("Notes")) or ""

        if not regatta or not season_year or not race_name:
            continue  # blank/instruction row
        if "example row" in notes.lower():
            skipped += 1
            continue
        if not norm(sail_no):
            continue

        try:
            season_year = int(season_year)
        except (ValueError, TypeError):
            print(f"  Skipping row - bad SeasonYear: {season_year!r}")
            continue

        key = (regatta, season_year, race_name)
        if key not in race_cache:
            regatta_id = get_or_create_regatta(cur, regatta, args.category)
            src = norm(get("SourceURL"))
            event_id = get_or_create_event(cur, regatta_id, season_year,
                                            source_url=src, notes="Loaded from manual entry template")
            race_id = get_or_create_race(cur, event_id, race_name)
            race_cache[key] = race_id
            n_races.add(key)
        race_id = race_cache[key]

        boat_name = get("BoatName")
        boat_type = get("BoatType")
        owner_name = get("Owner")
        skipper = get("Skipper") or owner_name
        sailmaker_raw = get("Sailmaker")
        position = get("Position")
        try:
            position = int(float(position)) if norm(position) else None
        except (ValueError, TypeError):
            position = None
        points = get("Points")
        try:
            points = float(points) if norm(points) else None
        except (ValueError, TypeError):
            points = None
        corrected = norm(get("CorrectedTime"))
        status = status_from(get("Status"), position, corrected)

        boat_id = get_or_create_boat(cur, sail_no, boat_name, boat_type, None)
        owner_id = get_or_create_owner(cur, owner_name)
        sm_id = get_or_create_sailmaker(cur, sailmaker_raw)

        cur.execute(
            "INSERT OR REPLACE INTO race_entries "
            "(id, race_id, boat_id, class, sail_no_used, boat_name_used, boat_type_used, "
            " owner_id, owner_name_used, skipper_name_used, sailmaker_id, status, "
            " finish_time, elapsed_time, corrected_time, position, points, comments, source) VALUES ("
            " (SELECT id FROM race_entries WHERE race_id = ? AND boat_id = ?),"
            " ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'manual:template')",
            (race_id, boat_id,
             race_id, boat_id, norm(get("Class")), norm_upper(sail_no), norm_upper(boat_name),
             norm(boat_type), owner_id, norm_upper(owner_name), norm_upper(skipper), sm_id, status,
             norm(get("FinishTime")), norm(get("ElapsedTime")), corrected, position, points,
             norm(get("Notes"))),
        )
        n_rows += 1

    conn.commit()
    print(f"Loaded {n_rows} entries across {len(n_races)} race(s)"
          + (f" (skipped {skipped} example row)" if skipped else "") + ".")
    for regatta, year, race_name in sorted(n_races):
        print(f"  - {regatta} {year}: {race_name}")
    conn.close()


if __name__ == "__main__":
    main()
